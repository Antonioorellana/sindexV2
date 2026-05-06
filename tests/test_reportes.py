import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from db.database import get_connection, initialize_database
from core.reportes import (
    exportar_descuentos_para_funs,
    exportar_historial_excel,
    obtener_datos_historial,
)

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


def _crear_db_historial(base_dir: Path) -> Path:
    db_path = base_dir / "sindicato.db"
    initialize_database(db_path)
    with get_connection(db_path) as connection:
        connection.executemany(
            "INSERT INTO socios (rut, nombre, local, activo, mes_carga) VALUES (?, ?, ?, ?, ?)",
            [
                ("15375441-1", "SOCIO UNO", "LOCAL", 1, "2026-05"),
                ("11111111-1", "SOCIO DOS", "LOCAL", 1, "2026-05"),
            ],
        )
        connection.executemany(
            "INSERT INTO descuentos_mensuales (rut, mes, tipo, monto, descripcion) VALUES (?, ?, ?, ?, ?)",
            [
                ("15375441-1", "2026-01", "Optica", 10000, ""),
                ("15375441-1", "2026-01", "Clinica", 5000, ""),
                ("15375441-1", "2026-03", "Otro", 7000, ""),
            ],
        )
        connection.executemany(
            "INSERT INTO cesjun_descuentos (mes, rut, nombre, monto) VALUES (?, ?, ?, ?)",
            [
                ("2026-01", "15375441-1", "SOCIO UNO", 12000),
                ("2026-03", "15375441-1", "SOCIO UNO", 7000),
            ],
        )
        connection.execute(
            """
            INSERT INTO conciliacion_mensual (
                mes, rut, nombre, total_calculado, total_descontado, diferencia, estado, observacion
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("2026-03", "15375441-1", "SOCIO UNO", 7000, 0, 7000, "no_descontado", ""),
        )
    return db_path


class ReportesTests(unittest.TestCase):
    def test_historial_anual_tiene_12_meses_y_sin_descuento(self):
        with TemporaryDirectory() as tmp:
            db_path = _crear_db_historial(Path(tmp))
            meses, filas = obtener_datos_historial(db_path, anio=2026)

        self.assertEqual(len(meses), 12)
        self.assertEqual(meses[0], "2026-01")
        self.assertEqual(meses[-1], "2026-12")
        socio_uno = next(f for f in filas if f["rut"] == "15375441-1")
        socio_dos = next(f for f in filas if f["rut"] == "11111111-1")
        self.assertEqual(socio_uno["2026-01"], 12000)
        self.assertIsNone(socio_dos.get("2026-01"))
        self.assertTrue(socio_uno["_problema"])

    @unittest.skipIf(load_workbook is None, "openpyxl no disponible")
    def test_exportar_funs_consolida_una_fila_por_socio(self):
        with TemporaryDirectory() as tmp:
            db_path = _crear_db_historial(Path(tmp))
            salida = Path(tmp) / "funs.xlsx"
            exportar_descuentos_para_funs("2026-01", salida, db_path)
            workbook = load_workbook(salida)
            hoja = workbook["Descuentos FUNS"]

        self.assertEqual([hoja["A2"].value, hoja["B2"].value, hoja["C2"].value], ["RUT", "Nombre", "Total descuento"])
        self.assertEqual(hoja["A3"].value, "15375441-1")
        self.assertEqual(hoja["C3"].value, 15000)
        self.assertEqual(hoja["B4"].value, "TOTAL GENERAL")
        self.assertEqual(hoja["C4"].value, 15000)

    @unittest.skipIf(load_workbook is None, "openpyxl no disponible")
    def test_exportar_historial_excel_muestra_sin_descuento(self):
        with TemporaryDirectory() as tmp:
            db_path = _crear_db_historial(Path(tmp))
            salida = Path(tmp) / "historial.xlsx"
            exportar_historial_excel(salida, db_path, anio=2026)
            workbook = load_workbook(salida)
            hoja = workbook["Historial"]

        self.assertEqual(hoja["C2"].value, "2026-01")
        self.assertEqual(hoja["N2"].value, "2026-12")
        fila_socio_uno = next(row for row in range(3, hoja.max_row + 1) if hoja.cell(row=row, column=1).value == "15375441-1")
        self.assertEqual(hoja.cell(row=fila_socio_uno, column=3).value, 12000)
        self.assertEqual(hoja.cell(row=fila_socio_uno, column=4).value, "sin descuento")


if __name__ == "__main__":
    unittest.main()
