from __future__ import annotations

from datetime import date
from pathlib import Path

from core.conciliacion import obtener_casos_por_estado, resumen_conciliacion
from db.database import get_connection

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - depende del entorno
    Workbook = None
    Font = PatternFill = Alignment = Border = Side = get_column_letter = None


def obtener_detalle_conciliacion(mes: str, db_path=None) -> list[dict[str, object]]:
    with get_connection(db_path) as connection:
        rows = connection.execute(
            """
            SELECT rut, nombre, total_calculado, total_descontado, diferencia, estado, observacion
            FROM conciliacion_mensual
            WHERE mes = ?
            ORDER BY estado, rut
            """,
            (mes,),
        ).fetchall()
    return [dict(row) for row in rows]


def obtener_resumen_por_estado(mes: str, db_path=None) -> dict[str, int]:
    return resumen_conciliacion(mes, db_path)["estados"]


def exportar_reporte_conciliacion(mes: str, ruta_salida: str | Path, db_path=None) -> Path:
    if Workbook is None:
        raise RuntimeError("openpyxl no esta disponible en este entorno")
    detalle = obtener_detalle_conciliacion(mes, db_path)
    resumen = resumen_conciliacion(mes, db_path)
    workbook = Workbook()
    hoja_resumen = workbook.active
    hoja_resumen.title = "Resumen"
    hoja_resumen.append(["Mes", resumen["mes"]])
    hoja_resumen.append(["Total enviado FUNS", resumen["total_enviado_funs"]])
    hoja_resumen.append(["Total descontado CESJUN", resumen["total_descontado_cesjun"]])
    hoja_resumen.append(["Diferencia", resumen["diferencia"]])
    hoja_resumen.append([])
    hoja_resumen.append(["Estado", "Cantidad"])
    for estado, cantidad in resumen["estados"].items():
        hoja_resumen.append([estado, cantidad])
    hoja_detalle = workbook.create_sheet("Detalle")
    hoja_detalle.append(
        ["RUT", "Nombre", "Monto enviado FUNS", "Monto descontado CESJUN", "Diferencia", "Estado", "Observacion"]
    )
    for fila in detalle:
        hoja_detalle.append(
            [
                fila["rut"],
                fila["nombre"],
                fila["total_calculado"],
                fila["total_descontado"],
                fila["diferencia"],
                fila["estado"],
                fila["observacion"],
            ]
        )
    destino = Path(ruta_salida)
    destino.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destino)
    return destino


# ── helpers de estilo ────────────────────────────────────────────────────────

def _fill(hex_color: str):
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _thin_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _apply_header_row(hoja, row_idx: int, labels: list[str], bg: str = "1A5276", fg: str = "FFFFFF") -> None:
    for col_idx, label in enumerate(labels, start=1):
        cell = hoja.cell(row=row_idx, column=col_idx, value=label)
        cell.font = Font(bold=True, color=fg, size=11)
        cell.fill = _fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()


def _auto_width(hoja, min_width: int = 10, max_width: int = 60) -> None:
    for col_cells in hoja.columns:
        length = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        hoja.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, min_width), max_width)


# ── exportar descuentos para FUNS ────────────────────────────────────────────

def exportar_descuentos_para_funs(mes: str, ruta_salida: str | Path, db_path=None) -> Path:
    """Exporta una fila por socio con el total consolidado del mes."""
    if Workbook is None:
        raise RuntimeError("openpyxl no esta disponible en este entorno")

    with get_connection(db_path) as connection:
        rows = connection.execute(
            """
            SELECT dm.rut,
                   COALESCE(s.nombre, dm.rut) AS nombre,
                   SUM(dm.monto) AS total
            FROM descuentos_mensuales dm
            LEFT JOIN socios s ON dm.rut = s.rut
            WHERE dm.mes = ?
            GROUP BY dm.rut
            ORDER BY COALESCE(s.nombre, dm.rut), dm.rut
            """,
            (mes,),
        ).fetchall()

    workbook = Workbook()
    hoja = workbook.active
    hoja.title = "Descuentos FUNS"

    hoja.merge_cells("A1:C1")
    titulo = hoja["A1"]
    titulo.value = f"FUNS consolidado  —  Mes {mes}"
    titulo.font = Font(bold=True, size=14, color="FFFFFF")
    titulo.fill = _fill("1A5276")
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    hoja.row_dimensions[1].height = 28

    _apply_header_row(hoja, 2, ["RUT", "Nombre", "Total descuento"])
    hoja.row_dimensions[2].height = 20

    fila_actual = 3
    total_general = 0
    amarillo = "FEF9E7"

    for registro in rows:
        bg = "FFFFFF" if (fila_actual % 2 == 0) else "F2F3F4"
        hoja.cell(row=fila_actual, column=1, value=registro["rut"]).border = _thin_border()
        hoja.cell(row=fila_actual, column=2, value=registro["nombre"]).border = _thin_border()
        monto = int(registro["total"])
        monto_cell = hoja.cell(row=fila_actual, column=3, value=monto)
        monto_cell.border = _thin_border()
        monto_cell.number_format = '#,##0'
        for col in range(1, 4):
            hoja.cell(row=fila_actual, column=col).fill = _fill(bg)
        total_general += monto
        fila_actual += 1

    hoja.cell(row=fila_actual, column=1, value="")
    hoja.cell(row=fila_actual, column=2, value="")
    hoja.cell(row=fila_actual, column=2, value="TOTAL GENERAL")
    total_cell = hoja.cell(row=fila_actual, column=3, value=total_general)
    total_cell.number_format = '#,##0'
    for col in range(1, 4):
        c = hoja.cell(row=fila_actual, column=col)
        c.fill = _fill(amarillo)
        c.font = Font(bold=True, size=11)
        c.border = _thin_border()

    _auto_width(hoja)
    destino = Path(ruta_salida)
    destino.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destino)
    return destino


# ── historial de descuentos (matriz) ─────────────────────────────────────────

MESES_ANUALES = [f"{mes:02d}" for mes in range(1, 13)]


def _normalizar_anio(anio: int | str | None) -> int:
    if anio in (None, ""):
        return date.today().year
    return int(str(anio).strip())


def obtener_datos_historial(
    db_path=None,
    rut_filtro: str | None = None,
    anio: int | str | None = None,
) -> tuple[list[str], list[dict[str, object]]]:
    """
    Retorna (meses, filas) donde:
    - meses: lista fija del año (ej. ["2026-01", ..., "2026-12"])
    - filas: lista de dicts con rut, nombre, un key por mes con el total o None,
             y "_problema": True si en algún mes el estado fue no_descontado/diferencia
    """
    anio_int = _normalizar_anio(anio)
    meses = [f"{anio_int}-{mes}" for mes in MESES_ANUALES]
    inicio = meses[0]
    fin = meses[-1]

    with get_connection(db_path) as connection:
        socios_params: list[object] = []
        socios_where = ""
        if rut_filtro:
            socios_where = "WHERE rut = ?"
            socios_params.append(rut_filtro)
        socio_rows = connection.execute(
            f"""
            SELECT rut, nombre
            FROM socios
            {socios_where}
            ORDER BY nombre, rut
            """,
            socios_params,
        ).fetchall()

        cesjun_conditions = ["cd.mes BETWEEN ? AND ?"]
        cesjun_params: list[object] = [inicio, fin]
        if rut_filtro:
            cesjun_conditions.append("cd.rut = ?")
            cesjun_params.append(rut_filtro)
        cesjun_rows = connection.execute(
            f"""
            SELECT cd.rut,
                   COALESCE(s.nombre, cd.nombre, cd.rut) AS nombre,
                   cd.mes,
                   SUM(cd.monto) AS total
            FROM cesjun_descuentos cd
            LEFT JOIN socios s ON cd.rut = s.rut
            WHERE {" AND ".join(cesjun_conditions)}
            GROUP BY cd.rut, cd.mes
            ORDER BY nombre, cd.mes
            """,
            cesjun_params,
        ).fetchall()

        conc_conditions = ["mes BETWEEN ? AND ?"]
        conc_params: list[object] = [inicio, fin]
        if rut_filtro:
            conc_conditions.append("rut = ?")
            conc_params.append(rut_filtro)
        conc_rows = connection.execute(
            f"SELECT rut, mes, estado FROM conciliacion_mensual WHERE {' AND '.join(conc_conditions)}",
            conc_params,
        ).fetchall()

    estados: dict[tuple[str, str], str] = {(r["rut"], r["mes"]): r["estado"] for r in conc_rows}

    ruts_data: dict[str, dict[str, object]] = {}
    for row in socio_rows:
        ruts_data[row["rut"]] = {"rut": row["rut"], "nombre": row["nombre"], "_problema": False}

    for row in cesjun_rows:
        rut = row["rut"]
        mes = row["mes"]
        if rut not in ruts_data:
            ruts_data[rut] = {"rut": rut, "nombre": row["nombre"], "_problema": False}
        ruts_data[rut][mes] = int(row["total"])
        estado = estados.get((rut, mes))
        if estado in ("no_descontado", "diferencia"):
            ruts_data[rut]["_problema"] = True

    filas = list(ruts_data.values())
    return meses, filas


def exportar_historial_excel(
    ruta_salida: str | Path,
    db_path=None,
    rut_filtro: str | None = None,
    anio: int | str | None = None,
) -> Path:
    """Exporta la matriz anual de descuentos a Excel con celdas en rojo para casos problemáticos."""
    if Workbook is None:
        raise RuntimeError("openpyxl no esta disponible en este entorno")

    anio_int = _normalizar_anio(anio)
    meses, filas = obtener_datos_historial(db_path, rut_filtro, anio_int)

    workbook = Workbook()
    hoja = workbook.active
    hoja.title = "Historial"

    n_cols = 2 + len(meses)

    # Título
    hoja.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(n_cols, 3))
    titulo = hoja["A1"]
    titulo.value = f"Historial anual de descuentos {anio_int}"
    if rut_filtro:
        titulo.value += f"  —  RUT {rut_filtro}"
    titulo.font = Font(bold=True, size=14, color="FFFFFF")
    titulo.fill = _fill("1A5276")
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    hoja.row_dimensions[1].height = 28

    # Encabezados
    labels = ["RUT", "Nombre"] + meses
    _apply_header_row(hoja, 2, labels)
    hoja.row_dimensions[2].height = 20

    rojo_fill = _fill("FADBD8")
    rojo_font_bold = Font(bold=True, color="922B21", size=10)
    rojo_font = Font(color="922B21", size=10)
    normal_font = Font(size=10)
    normal_font_bold = Font(bold=True, size=10)

    for fila_idx, fila in enumerate(filas, start=3):
        es_problema = bool(fila.get("_problema"))
        bg = "FADBD8" if es_problema else ("FFFFFF" if fila_idx % 2 == 1 else "F2F3F4")
        hoja.row_dimensions[fila_idx].height = 16

        c_rut = hoja.cell(row=fila_idx, column=1, value=fila["rut"])
        c_nom = hoja.cell(row=fila_idx, column=2, value=fila["nombre"])
        c_rut.font = rojo_font_bold if es_problema else normal_font_bold
        c_nom.font = rojo_font_bold if es_problema else normal_font_bold

        for col_idx, mes in enumerate(meses, start=3):
            total = fila.get(mes)
            cell = hoja.cell(row=fila_idx, column=col_idx)
            if total is not None:
                cell.value = int(total)
                cell.number_format = '#,##0'
            else:
                cell.value = "sin descuento"
            cell.font = rojo_font if es_problema else normal_font
            cell.fill = _fill(bg)
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="right" if total is not None else "center")

        for col in (1, 2):
            c = hoja.cell(row=fila_idx, column=col)
            c.fill = _fill(bg)
            c.border = _thin_border()

    _auto_width(hoja)
    destino = Path(ruta_salida)
    destino.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destino)
    return destino
