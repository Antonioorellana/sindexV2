from __future__ import annotations

from numbers import Number
from pathlib import Path

from core.ruts import preparar_rut_para_busqueda

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


def _require_openpyxl() -> None:
    if load_workbook is None:
        raise RuntimeError("openpyxl no esta disponible en este entorno")


def _normalizar_encabezado(valor: object) -> str:
    return str(valor or "").strip().lower()


def _entero(valor: object) -> int:
    if valor is None:
        return 0
    if isinstance(valor, Number):
        return int(valor)
    texto = str(valor or "").strip().replace("$", "").replace(" ", "")
    if "," in texto:
        texto = texto.split(",", 1)[0]
    elif texto.count(".") == 1 and len(texto.rsplit(".", 1)[1]) <= 2:
        texto = texto.rsplit(".", 1)[0]
    texto = texto.replace(".", "").replace(",", "")
    try:
        return int(texto)
    except ValueError:
        return 0


def _encontrar_fila_encabezado(hoja, max_busqueda: int = 20) -> tuple[int, dict[str, int]]:
    """Busca la fila que contiene los encabezados detectando la columna 'rut'."""
    for row_idx in range(1, max_busqueda + 1):
        fila = list(hoja.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        encabezados: dict[str, int] = {}
        for col_idx, valor in enumerate(fila, start=1):
            key = _normalizar_encabezado(valor)
            if key:
                encabezados[key] = col_idx
        if "rut" in encabezados:
            return row_idx, encabezados
    return 1, {}


def _importar_hoja_sijuan(hoja) -> list[dict[str, object]]:
    header_row, encabezados = _encontrar_fila_encabezado(hoja)
    socios = []
    for fila in hoja.iter_rows(min_row=header_row + 1, values_only=True):
        rut_valor = fila[encabezados.get("rut", 0) - 1] if encabezados.get("rut") else None
        nombre = fila[encabezados.get("nombre", 0) - 1] if encabezados.get("nombre") else None
        if not rut_valor or not nombre:
            continue
        try:
            rut = preparar_rut_para_busqueda(str(rut_valor))
        except ValueError:
            continue
        socios.append(
            {
                "rut": rut,
                "nombre": str(nombre).strip(),
                "local": str(fila[encabezados.get("local", 0) - 1] or "").strip()
                if encabezados.get("local")
                else "",
            }
        )
    return socios


def _importar_hoja_cesjun(hoja) -> list[dict[str, object]]:
    header_row, encabezados = _encontrar_fila_encabezado(hoja)
    registros = []
    for fila in hoja.iter_rows(min_row=header_row + 1, values_only=True):
        rut_valor = fila[encabezados.get("rut", 0) - 1] if encabezados.get("rut") else None
        monto_valor = fila[encabezados.get("monto", 0) - 1] if encabezados.get("monto") else None
        if not rut_valor or monto_valor in (None, ""):
            continue
        try:
            rut = preparar_rut_para_busqueda(str(rut_valor))
        except ValueError:
            continue
        registros.append(
            {
                "rut": rut,
                "nombre": str(fila[encabezados.get("nombre", 0) - 1] or "").strip()
                if encabezados.get("nombre")
                else "",
                "monto": _entero(monto_valor),
            }
        )
    return registros


def importar_archivo_mensual(ruta_archivo: str | Path) -> dict[str, list[dict[str, object]]]:
    """Lee el archivo mensual con hojas SIJUAN (activos) y CESJUN (descontado)."""
    _require_openpyxl()
    workbook = load_workbook(ruta_archivo, data_only=False)
    names_map = {name.lower(): name for name in workbook.sheetnames}

    sijuan_name = names_map.get("sijuan") or workbook.sheetnames[0]
    cesjun_name = next(
        (names_map[k] for k in names_map if "cesjun" in k),
        workbook.sheetnames[1] if len(workbook.sheetnames) > 1 else None,
    )

    socios = _importar_hoja_sijuan(workbook[sijuan_name])
    cesjun = _importar_hoja_cesjun(workbook[cesjun_name]) if cesjun_name else []
    return {"socios": socios, "cesjun": cesjun}


def importar_sijuan(ruta_archivo: str | Path) -> list[dict[str, object]]:
    _require_openpyxl()
    workbook = load_workbook(ruta_archivo, data_only=False)
    # Si el archivo tiene hoja SIJUAN, usarla; si no, usar hoja activa
    names_map = {name.lower(): name for name in workbook.sheetnames}
    hoja_name = names_map.get("sijuan") or workbook.active.title
    return _importar_hoja_sijuan(workbook[hoja_name])


def importar_cesjun(ruta_archivo: str | Path) -> list[dict[str, object]]:
    _require_openpyxl()
    workbook = load_workbook(ruta_archivo, data_only=False)
    names_map = {name.lower(): name for name in workbook.sheetnames}
    hoja_name = next(
        (names_map[k] for k in names_map if "cesjun" in k),
        workbook.active.title,
    )
    return _importar_hoja_cesjun(workbook[hoja_name])


def importar_funs_enviado(ruta_archivo: str | Path) -> list[dict[str, object]]:
    """Lee el FUNS (plantilla empresa). Columnas fijas: H=RUT, I=nombre, J=motivo, K=monto."""
    _require_openpyxl()
    workbook = load_workbook(ruta_archivo, data_only=False)
    hoja = workbook["Descuentos"] if "Descuentos" in workbook.sheetnames else workbook.active
    registros = []
    for fila in hoja.iter_rows(min_row=8, values_only=True):
        rut_valor = fila[7]   # columna H
        nombre    = fila[8]   # columna I
        motivo    = fila[9]   # columna J
        monto_valor = fila[10]  # columna K
        # Saltar formulas, encabezados y filas vacías
        if not rut_valor or monto_valor in (None, ""):
            continue
        rut_str = str(rut_valor).strip()
        if rut_str.startswith("=") or not rut_str:
            continue
        try:
            rut = preparar_rut_para_busqueda(rut_str)
        except ValueError:
            continue
        monto = _entero(monto_valor)
        if monto <= 0:
            continue
        registros.append(
            {
                "rut": rut,
                "nombre": str(nombre or "").strip(),
                "motivo": str(motivo or "").strip(),
                "monto": monto,
            }
        )
    return registros
